# -*- coding: utf-8 -*-
"""
Created on Wed Sep 27 11:34:32 2023

@author: brian
"""

import openpyxl
from aircraft_categories import AircraftCatalog, Aircraft


def update_worksheet(sheet):
    for row_number, cell in enumerate(sheet.iter_rows(min_row=5, min_col=3, max_col=3, values_only=True), start=5):
        if cell:
            aircraft_type = cell[0]
            aircraft_category = catalog.get_category_by_aircraft(aircraft_type)
            aircraft_class = catalog.get_class_by_aircraft(aircraft_type)
            
            aircraft = Aircraft(aircraft_type, aircraft_class, aircraft_category)
            print(f"Name: {aircraft.name}, Class: {aircraft.classification}, Category: {aircraft.category}")
            # Print the category & classes in the corresponding cell 
            sheet.cell(row=row_number, column=5, value=aircraft_class)
            sheet.cell(row=row_number, column=6, value=aircraft_category)
    
    file_path = ('/home/lembuss/Desktop/Semester Thesis/Project Documents/Traffic_26_09_EDDM_updated.xlsx')
    traffic_data_workbook.save(file_path)
    

# Load the Excel spreadsheet
file_path = ('/home/lembuss/Desktop/Semester Thesis/Project Documents/Traffic_26_09_EDDM.xlsx')
traffic_data_workbook = openpyxl.load_workbook(file_path)  


# Create an AircraftCatalog and populate it with data from the spreadsheet
aircraft_category_sheet = traffic_data_workbook['RECAT_EU']
catalog = AircraftCatalog()


# first row contains classification, second row category letters and the aircraft types are listed in columns under each category
for col in aircraft_category_sheet.iter_cols(min_row=1, max_row=100, values_only=True):
    aircraft_class = col[0]
    aircraft_category = col[1]
    for cell in col[2:]:
        if cell:
            catalog.add_aircraft(cell, aircraft_class, aircraft_category)

print("All Aircraft:")
catalog.display_all_aircraft()

print("Now let us work on updating the sheets")
# use the catalog to sort aircraft categories in the table
traffic_sheets = ["Arrivals - Aircraft Type","Departures - Aircraft Type"]
# aircraft_arrivals_sheet = traffic_data_workbook['Arrivals - Aircraft Type']
# aircraft_departures_sheet = traffic_data_workbook['Departures - Aircraft Type']

for sheet_name in traffic_sheets:
    sheet = traffic_data_workbook[sheet_name]
    update_worksheet(sheet)
    
    



