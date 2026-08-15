import openpyxl
from aircraft_categories import AircraftCatalog
from glideslope_visualization import starting_distance_ft, x_at_1000ft_separation, x_at_evtol_cruising_altitude, convert_ft_to_nm
import simplekml
from datetime import timedelta, datetime, time, date
import geopandas as gpd
import math

def update_catalog(catalog):
    # first row contains classification, second row category letters and the aircraft types are listed in columns under each category
    for col in aircraft_category_sheet.iter_cols(min_row=1, max_row=100, values_only=True):
        aircraft_class = col[0]
        aircraft_category = col[1]
        for cell in col[2:]:
            if cell:
                catalog.add_aircraft(cell, aircraft_class, aircraft_category)

def update_dataset(traffic_sheets, aircraft_dataset):
    # access aircraft catalog & populate dataset
    for row in traffic_sheets.iter_rows(min_row=2, max_row=500, values_only=True):
        arrival_time = row[0]
        aircraft_name = row[4]
        aircraft_category = catalog.get_category_by_aircraft(str(aircraft_name))
        aircraft_dataset.append({'name': aircraft_name, 'category': aircraft_category, 'arrival_time': arrival_time})
        #print(f"Name: {aircraft_type}, Arrival Time: {arrival_time}, Category: {aircraft_category}")

def create_master_kml(aircraft_dataset, category_to_kml, output_filename):
    master_kml = simplekml.Kml()

    limit = 10
    counter = 0
    for aircraft in aircraft_dataset:

        if not aircraft['name'] or not aircraft['arrival_time'] or not aircraft['category']:
            continue
        # Create a NetworkLink
        simulation_date = date(2023, 9, 26)
        netlink = master_kml.newnetworklink(name=f"Aircraft Arrival at {aircraft['arrival_time']}")

        # Set the TimeSpan considering the glideslope considerations
        arrival_datetime = datetime.combine(simulation_date, aircraft['arrival_time'])
        start_delta, end_delta = restrictedarea_triggers()
        
        #netlink.timespan.begin = (arrival_datetime - timedelta(minutes=10)).isoformat()
        netlink.timespan.begin = (arrival_datetime - timedelta(minutes=start_delta)).isoformat()
        netlink.timespan.end = (arrival_datetime - timedelta(minutes=end_delta)).isoformat()
        
        #netlink.timespan.end = arrival_datetime.isoformat()

        # Set the link to the corresponding KML file based on aircraft category
        netlink.link.href = category_to_kml[aircraft['category']]
        counter += 1
        if counter == limit:
            break
    
    print(counter)
    # Save the KML file
    master_kml.save(output_filename)

def restrictedarea_triggers ():
    cruisespeed = 150 # speed in nautical miles per hour
    approach_duration = convert_ft_to_nm(starting_distance_ft)/cruisespeed
    
    start_distance = convert_ft_to_nm(starting_distance_ft - x_at_1000ft_separation)
    start_time = start_distance/cruisespeed

    end_distance = convert_ft_to_nm(starting_distance_ft - x_at_evtol_cruising_altitude)
    end_time = end_distance/cruisespeed

    start_delta =(approach_duration - start_time) * 60
    end_delta = (approach_duration -end_time) * 60

    return math.ceil(start_delta), math.ceil(end_delta)

parent_filepath = '/home/lembuss/Desktop/Semester Thesis/Dynamic_Airspace_Reconfiguration'

# Load the Excel spreadsheet
file_path = parent_filepath + '/Input_Data/Traffic_26_09_EDDM_updated.xlsx'
traffic_data_workbook = openpyxl.load_workbook(file_path)  

# Create an AircraftCatalog and populate it with data from the spreadsheet
aircraft_category_sheet = traffic_data_workbook['RECAT_EU']
catalog = AircraftCatalog()
update_catalog(catalog)

# Assign traffic to correct category and class and fill arrival time in a dataset
traffic_sheets = traffic_data_workbook['Arrivals']
aircraft_dataset = []
update_dataset(traffic_sheets, aircraft_dataset)
        

#print(aircraft_dataset)

category_to_kml = {
    'CAT-A': '/home/lembuss/Desktop/Semester Thesis/Dynamic_Airspace_Reconfiguration/Restricted_Areas/cata_apch.kml',
    'CAT-B': '/home/lembuss/Desktop/Semester Thesis/Dynamic_Airspace_Reconfiguration/Restricted_Areas/catb_apch.kml',
    'CAT-C': '/home/lembuss/Desktop/Semester Thesis/Dynamic_Airspace_Reconfiguration/Restricted_Areas/catc_apch.kml',
    'CAT-D': '/home/lembuss/Desktop/Semester Thesis/Dynamic_Airspace_Reconfiguration/Restricted_Areas/catd_apch.kml',
    'CAT-E': '/home/lembuss/Desktop/Semester Thesis/Dynamic_Airspace_Reconfiguration/Restricted_Areas/cate_apch.kml',
    'CAT-F': '/home/lembuss/Desktop/Semester Thesis/Dynamic_Airspace_Reconfiguration/Restricted_Areas/catf_apch.kml'
}

output_filename = "full_sim_flights.kml"
output_filepath = parent_filepath + '/Output_Map_Files/' + output_filename

# create simulation kml
create_master_kml(aircraft_dataset, category_to_kml, output_filepath)

# print(aircraft_dataset)   




